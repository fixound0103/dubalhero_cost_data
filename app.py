import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import re
from datetime import datetime

# 1. 웹페이지 기본 설정
st.set_page_config(page_title="체인로지스 정산 프로그램", layout="centered")

st.title("📊 체인로지스 정산 데이터 변환 프로그램")
st.write("원본 정산양식 엑셀 파일을 업로드하고 '분석 시작' 버튼을 클릭하세요.")

# 2. 파일 업로드 (xlsx 전용)
uploaded_file = st.file_uploader("정산양식 엑셀 파일(xlsx)을 업로드해주세요.", type=["xlsx"])

if uploaded_file is not None:
    st.success(f"파일 업로드 완료: {uploaded_file.name}")

    # 3. 분석 시작 버튼
    if st.button("분석 시작"):
        try:
            with st.spinner('원본 코드와 동일하게 데이터를 정밀 집계 중입니다...'):
                
                # --- [동적 파일명 및 월 추출 로직] ---
                input_filename = uploaded_file.name
                month_match = re.search(r'(\d+월)', input_filename)
                
                if month_match:
                    extracted_month = month_match.group(1)  # 예: "5월"
                else:
                    extracted_month = f"{datetime.now().month}월"  # 없을 경우 현재 월 기본값
                
                download_filename = f"체인로지스 정산_{extracted_month}.xlsx"
                sheet_title = f"{extracted_month} 정산결과"
                accident_title = f"{extracted_month} 사고내역 0건"

                # --- [중요: Streamlit 파일 포인터 방지용 버퍼 복사] ---
                # 하나의 파일을 두 번(야간, 주간) 읽을 때 발생하는 데이터 유실을 방지합니다.
                file_bytes = uploaded_file.read()

                # 4. 시트별 데이터 전처리 및 집계 함수 (원본 로직 100% 일치)
                def process_sheet_data(data_bytes, sheet_name):
                    df = pd.read_excel(BytesIO(data_bytes), sheet_name=sheet_name, engine='openpyxl')
                    df.columns = df.columns.str.strip()

                    # 날짜 정제 및 통합
                    df["접수일자"] = pd.to_datetime(df["배송완료"], errors="coerce").dt.strftime("%Y-%m-%d")
                    df = df.dropna(subset=["접수일자"])

                    # 고유 ID 컬럼 자동 매칭
                    id_col = [c for c in df.columns if "번호" in c or "ID" in c or "운송장" in c][0]

                    summary = (
                        df.groupby("접수일자")
                        .agg(배송건=(id_col, "nunique"), 단가=("요금", "median"))
                        .reset_index()
                    )
                    return summary

                # 야간(당일배송), 주간(회차배송) 탭 각각 가공 및 결합
                df_night = process_sheet_data(file_bytes, "야간")
                df_day = process_sheet_data(file_bytes, "주간")

                all_dates = pd.merge(
                    df_night, df_day, on="접수일자", how="outer", suffixes=("_당일", "_회차")
                ).fillna(0)
                all_dates = all_dates.sort_values(by="접수일자").reset_index(drop=True)

                # 5. 새로운 엑셀 워크북 생성 및 기본 양식 설정
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_title
                ws.views.sheetView[0].showGridLines = True

                # 2단 헤더 생성
                ws.merge_cells("A1:A2")
                ws.merge_cells("B1:E1")
                ws.merge_cells("F1:I1")
                ws.merge_cells("J1:K1")

                ws["A1"] = "접수일자"
                ws["B1"] = "당일배송"
                ws["F1"] = "회차배송"
                ws["J1"] = "총합계"

                headers_level2 = [
                    "단가", "배송건", "금 액", "부가세",
                    "단가", "배송건", "금 액", "부가세",
                    "배송건수", "금 액",
                ]
                for col_idx, header in enumerate(headers_level2, start=2):
                    ws.cell(row=2, column=col_idx, value=header)

                # 6. 일자별 데이터 기록 및 기본 계산 수식 주입 (원본 수식 구조 100% 일치)
                pickup_count = 0
                start_row = 3

                for idx, row in all_dates.iterrows():
                    r = start_row + idx
                    ws.cell(row=r, column=1, value=row["접수일자"])

                    day_cnt = int(row["배송건_당일"])
                    cycle_cnt = int(row["배송건_회차"])

                    # ★ [원본 중요 조건] 총 배송건수가 10건 이상인 날만 픽업비용 횟수로 카운트
                    if (day_cnt + cycle_cnt) >= 10:
                        pickup_count += 1

                    # 당일배송
                    ws.cell(row=r, column=2, value=row["단가_당일"] if row["단가_당일"] > 0 else 2400)
                    ws.cell(row=r, column=3, value=day_cnt)
                    ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
                    ws.cell(row=r, column=5, value=f"=D{r}*0.1")

                    # 회차배송
                    ws.cell(row=r, column=6, value=row["단가_회차"] if row["단가_회차"] > 0 else 3500)
                    ws.cell(row=r, column=7, value=cycle_cnt)
                    ws.cell(row=r, column=8, value=f"=F{r}*G{r}")
                    ws.cell(row=r, column=9, value=f"=H{r}*0.1")

                    # 총합계 열
                    ws.cell(row=r, column=10, value=f"=C{r}+G{r}")
                    ws.cell(row=r, column=11, value=f"=SUM(D{r}:E{r})+SUM(H{r}:I{r})")

                data_end_row = start_row + len(all_dates) - 1

                # 7. 하단 특수 행 추가 (픽업비용 조건부 반영)
                
                # (1) 픽업비용 행 추가
                pickup_row = data_end_row + 1
                ws.merge_cells(f"A{pickup_row}:F{pickup_row}")
                ws.cell(row=pickup_row, column=1, value=f"픽업비용({pickup_count}회)")
                ws.cell(row=pickup_row, column=10, value=pickup_count)
                ws.cell(row=pickup_row, column=8, value=f"=60000*J{pickup_row}")
                ws.cell(row=pickup_row, column=9, value=f"=H{pickup_row}*0.1")
                ws.cell(row=pickup_row, column=11, value=f"=SUM(H{pickup_row}:I{pickup_row})")

                # (2) 사고내역 행 추가
                accident_row = pickup_row + 1
                ws.merge_cells(f"A{accident_row}:J{accident_row}")
                ws.cell(row=accident_row, column=1, value=accident_title)
                ws.cell(row=accident_row, column=11, value=0)

                # (3) 최종 [총합계] 행 추가 및 세로 합계 수식 연동
                total_row = accident_row + 1
                ws.cell(row=total_row, column=1, value="총합계")

                # 당일배송 세로 합계
                ws.cell(row=total_row, column=3, value=f"=SUM(C3:C{data_end_row})")
                ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{data_end_row})")
                ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{data_end_row})")

                # 회차배송 세로 합계
                ws.cell(row=total_row, column=7, value=f"=SUM(G3:G{data_end_row})")
                ws.cell(row=total_row, column=8, value=f"=SUM(H3:H{pickup_row})")
                ws.cell(row=total_row, column=9, value=f"=SUM(I3:I{pickup_row})")

                # 총합계 배송건수 세로 합산
                ws.cell(row=total_row, column=10, value=f"=SUM(J3:J{data_end_row})")

                # 최종 금액 정산 수식
                ws.cell(row=total_row, column=11, value=f"=SUM(K3:K{pickup_row})+K{accident_row}")

                # 8. 디자인 스타일 및 포맷 적용 (폰트, 테두리, 정렬 백분의 일 일치)
                header_fill = PatternFill(start_color="8FCE00", end_color="8FCE00", fill_type="solid")
                font_header = Font(name="맑은 고딕", size=11, bold=True)
                font_data = Font(name="맑은 고딕", size=11)
                font_total = Font(name="맑은 고딕", size=11, bold=True)

                thin_border = Border(
                    left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000")
                )

                for r in range(1, total_row + 1):
                    for c in range(1, 12):
                        cell = ws.cell(row=r, column=c)
                        cell.border = thin_border

                        if r in [1, 2]:
                            cell.fill = header_fill
                            cell.font = font_header
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif r == total_row:
                            cell.font = font_total
                            cell.alignment = Alignment(horizontal="center", vertical="center") if c == 1 else Alignment(horizontal="right", vertical="center")
                            if c > 1: cell.number_format = "#,##0"
                        elif r in [pickup_row, accident_row]:
                            cell.font = font_data
                            cell.alignment = Alignment(horizontal="center", vertical="center") if c == 1 else Alignment(horizontal="right", vertical="center")
                            if c > 1: cell.number_format = "#,##0"
                        else:
                            cell.font = font_data
                            cell.alignment = Alignment(horizontal="center", vertical="center") if c == 1 else Alignment(horizontal="right", vertical="center")
                            if c > 1: cell.number_format = "#,##0"

                # 9. 열 너비 자동 맞춤
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                # --- 물리 파일 대신 메모리 스트림을 활용해 즉시 다운로드 제공 ---
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                processed_data = excel_buffer.getvalue()

                st.balloons()
                st.success("✨ 원본 분석 프로그램과 100% 동일하게 변환 완료!")

                # 10. 순수 xlsx 다운로드 버튼 활성화
                st.download_button(
                    label="📥 변환된 정산 파일 다운로드",
                    data=processed_data,
                    file_name=download_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"⚠️ 에러 발생: {e}")
            st.info("업로드한 엑셀 파일에 원본과 동일한 명칭의 '야간' 및 '주간' 시트가 존재하는지 재차 확인해 주세요.")
